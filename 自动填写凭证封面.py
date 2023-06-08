from win32com.client import Dispatch
import openpyxl
import os
import sys
exepath = os.path.dirname(os.path.realpath(sys.argv[0]))
def typing(box,text):
    doc.Shapes.Range(box).Select()
    s.Text=text
def cleaning():
    typing("Text Box 3", "")#本月共有
    typing("Text Box 4", "")#当前册数
    typing("Text Box 5", "")  # 年份1
    typing("Text Box 8", "")  # 年份2
    typing("Text Box 6", "")  # 月份1
    typing("Text Box 9", "")  # 月份2
    typing("Text Box 10", "")  # 最后一日
    typing("Text Box 11", "")#利润中心box1
    typing("Text Box 12", "")#利润中心box2
    typing("Text Box 13", "")#利润中心box3
    typing("Text Box 14", "")#利润中心box4
    typing("Text Box 15", "")#月流水号自1
    typing("Text Box 16", "")#月流水号自2
    typing("Text Box 17", "")#月流水号自3
    typing("Text Box 18", "")#月流水号自4
    typing("Text Box 19", "")#月流水号至1
    typing("Text Box 20", "")#月流水号至2
    typing("Text Box 21", "")#月流水号至3
    typing("Text Box 22", "")#月流水号至4
    typing("Text Box 24", "")  # 利润中心组
def testsaving():
    if ws.cell(2 + i, 7).value != ws.cell(3 + i, 7).value:
        doc.SaveAs(exepath + "\\输出\\"+'凭证封面第' + str(ws.cell(2 + i, 7).value) + '本.docx')
        global centernum
        centernum = 1
        cleaning()
ws = openpyxl.load_workbook('凭证编号.xlsx')["Sheet1"]
path=exepath+"\\凭证封面打印版.docx"
app = Dispatch('Word.Application')
# 新建word文档
doc = app.Documents.Open(FileName=path)
app.Visible = False
s=app.Selection
cleaning()
centernum = 1  # 表示这本凭证里是第几个项目
for i in range(ws.max_row):
    try:
        typing("Text Box 3", ws.cell(2 + i, 3).value)
        typing("Text Box 4", ws.cell(2+i,7).value)
        typing("Text Box 5", ws.cell(2 + i, 8).value)
        typing("Text Box 8", ws.cell(2 + i, 8).value)
        typing("Text Box 6", ws.cell(2 + i, 9).value)
        typing("Text Box 9", ws.cell(2 + i, 9).value)
        typing("Text Box 10", ws.cell(2 + i, 10).value)
        typing("Text Box 2", ws.cell(2 + i, 1).value[1:5])
        typing("Text Box 24", ws.cell(2 + i, 11).value)
        if centernum==1:
            typing("Text Box 11", ws.cell(2+i,1).value+" "+ws.cell(2+i,2).value)
            typing("Text Box 15", ws.cell(2 + i, 4).value)
            typing("Text Box 19", ws.cell(2 + i, 5).value)
            centernum += 1
            print(ws.cell(2+i,1).value,centernum,ws.cell(2+i,7).value)
            testsaving()
            continue
        if centernum==2:
            typing("Text Box 12", ws.cell(2+i,1).value+" "+ws.cell(2+i,2).value)
            typing("Text Box 16", ws.cell(2 + i, 4).value)
            typing("Text Box 20", ws.cell(2 + i, 5).value)
            centernum += 1
            print(ws.cell(2+i,1).value,centernum,ws.cell(2+i,7).value)
            testsaving()
            continue
        if centernum==3:
            typing("Text Box 13", ws.cell(2 + i, 1).value + " " + ws.cell(2 + i, 2).value)
            typing("Text Box 17", ws.cell(2 + i, 4).value)
            typing("Text Box 21", ws.cell(2 + i, 5).value)
            centernum += 1
            print(ws.cell(2+i,1).value,centernum,ws.cell(2+i,7).value)
            testsaving()
            continue
        if centernum==4:
            typing("Text Box 14", ws.cell(2 + i, 1).value + " " + ws.cell(2 + i, 2).value)
            typing("Text Box 18", ws.cell(2 + i, 4).value)
            typing("Text Box 22", ws.cell(2 + i, 5).value)
            centernum += 1
            print(ws.cell(2+i,1).value,centernum,ws.cell(2+i,7).value)
            testsaving()
            continue

    except:
        pass
doc.Close()
app.Quit()
