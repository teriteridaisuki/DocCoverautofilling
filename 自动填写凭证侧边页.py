from win32com.client import Dispatch
import openpyxl


def typing(box, text):
    doc.Shapes.Range(box).Select()
    s.Text = text


def cleaning():
    typing("Text Box 1", "")#利润中心编码box1
    typing("Text Box 2", "")#利润中心编码box2
    typing("Text Box 7", "")#利润中心编码box3
    typing("Text Box 8", "")#利润中心编码box4
    typing("Text Box 11", "")#所在凭证册数
    typing("Text Box 12", "")  # 所有凭证册数
    typing("Text Box 13", "")#利润中心编码
    typing("Text Box 14", "")#月份
    typing("Text Box 15", "")#年份

def testsaving():
    if ws.cell(2 + i, 7).value != ws.cell(3 + i, 7).value:
        doc.SaveAs('E:\\untitle\\输出\\凭证侧边第' + str(ws.cell(2 + i, 7).value) + '本.docx')
        global centernum
        centernum = 1
        cleaning()


ws = openpyxl.load_workbook('凭证编号.xlsx')["Sheet1"]
path = r'E:\untitle\备用程序\凭证侧边与封面\凭证侧边打印版.docx'
app = Dispatch('Word.Application')
# 新建word文档
doc = app.Documents.Open(FileName=path)
app.Visible = False
s = app.Selection
cleaning()
centernum = 1  # 表示这本凭证里是第几个利润中心
for i in range(ws.max_row):
    try:
        typing("Text Box 1", ws.cell(2 + i, 1).value[1:5])#公司代码
        typing("Text Box 11", ws.cell(2 + i, 7).value)  # 所在凭证册数
        typing("Text Box 12", ws.cell(2 + i, 3).value)  # 所有凭证册数
        typing("Text Box 14", ws.cell(2 + i, 9).value)#月份
        typing("Text Box 15", ws.cell(2 + i, 8).value)  # 年份
        if centernum == 1:
            typing("Text Box 13", ws.cell(2 + i, 1).value)#利润中心编码
            centernum += 1
            print(ws.cell(2 + i, 1).value, centernum, ws.cell(2 + i, 7).value)
            testsaving()
            continue
        if centernum == 2:
            typing("Text Box 2", ws.cell(2 + i, 1).value)
            centernum += 1
            print(ws.cell(2 + i, 1).value, centernum, ws.cell(2 + i, 7).value)
            testsaving()
            continue
        if centernum == 3:
            typing("Text Box 7", ws.cell(2 + i, 1).value)
            centernum += 1
            print(ws.cell(2 + i, 1).value, centernum, ws.cell(2 + i, 7).value)
            testsaving()
            continue
        if centernum == 4:
            typing("Text Box 8", ws.cell(2 + i, 1).value)
            centernum += 1
            print(ws.cell(2 + i, 1).value, centernum, ws.cell(2 + i, 7).value)
            testsaving()
            continue

    except:
        pass
doc.Close()
app.Quit()

